import os
import subprocess
from datetime import datetime
from pathlib import Path
from app.core.config import settings
import shutil
import zipfile
import tarfile
from sqlalchemy.orm import Session
from sqlalchemy import inspect
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from typing import Dict, List, Any


class BackupService:
    def __init__(self):
        self.backup_dir = Path(settings.BACKUP_DIR if hasattr(settings, 'BACKUP_DIR') else './backups')
        self.backup_dir.mkdir(parents=True, exist_ok=True)
        self.max_backups = int(os.getenv('MAX_BACKUPS', '30'))  # Son 30 yedek sakla

    def create_database_backup(self, db_host: str, db_port: int, db_name: str, db_user: str, db_password: str) -> str:
        """Veritabanı yedeği oluştur"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_filename = f"backup_{db_name}_{timestamp}.sql"
        backup_path = self.backup_dir / backup_filename

        # pg_dump komutu
        env = os.environ.copy()
        env['PGPASSWORD'] = db_password

        cmd = [
            'pg_dump',
            '-h', db_host,
            '-p', str(db_port),
            '-U', db_user,
            '-d', db_name,
            '-F', 'c',  # Custom format
            '-f', str(backup_path)
        ]

        try:
            result = subprocess.run(cmd, env=env, capture_output=True, text=True, check=True)
            print(f"✅ Veritabanı yedeği oluşturuldu: {backup_path}")
            
            # Eski yedekleri temizle
            self.cleanup_old_backups()
            
            return str(backup_path)
        except subprocess.CalledProcessError as e:
            print(f"❌ Yedek oluşturma hatası: {e.stderr}")
            raise

    def cleanup_old_backups(self):
        """Eski yedekleri temizle (max_backups sayısından fazlasını sil)"""
        backups = sorted(self.backup_dir.glob('backup_*.sql'), key=os.path.getmtime, reverse=True)
        
        if len(backups) > self.max_backups:
            for backup in backups[self.max_backups:]:
                try:
                    backup.unlink()
                    print(f"🗑️ Eski yedek silindi: {backup.name}")
                except Exception as e:
                    print(f"⚠️ Yedek silme hatası: {e}")

    def list_backups(self) -> list:
        """Mevcut yedekleri listele"""
        backups = sorted(self.backup_dir.glob('backup_*.sql'), key=os.path.getmtime, reverse=True)
        return [
            {
                'filename': backup.name,
                'path': str(backup),
                'size': backup.stat().st_size,
                'created_at': datetime.fromtimestamp(backup.stat().st_mtime).isoformat()
            }
            for backup in backups
        ]

    def delete_backup(self, filename: str) -> bool:
        """Yedek dosyasını sil"""
        # Path traversal saldırısını önle
        if '..' in filename or '/' in filename or '\\' in filename:
            raise ValueError("Geçersiz dosya adı")
        
        # Dosya adını normalize et
        filename = os.path.basename(filename)
        
        backup_path = self.backup_dir / filename
        
        # Path traversal kontrolü (ikinci kez)
        real_path = os.path.realpath(str(backup_path))
        real_backup_dir = os.path.realpath(str(self.backup_dir))
        if not real_path.startswith(real_backup_dir):
            raise ValueError("Geçersiz dosya yolu")
        
        if not backup_path.exists():
            raise FileNotFoundError(f"Yedek dosyası bulunamadı: {filename}")
        
        try:
            backup_path.unlink()
            print(f"🗑️ Yedek silindi: {filename}")
            return True
        except Exception as e:
            print(f"⚠️ Yedek silme hatası: {e}")
            raise

    def restore_backup(self, backup_path: str, db_host: str, db_port: int, db_name: str, db_user: str, db_password: str):
        """Yedekten geri yükleme (DİKKAT: Bu işlem veritabanını sıfırlar!)"""
        env = os.environ.copy()
        env['PGPASSWORD'] = db_password

        # Önce mevcut veritabanını drop/create et (tehlikeli!)
        # Bu işlem sadece admin tarafından manuel olarak yapılmalı
        cmd = [
            'pg_restore',
            '-h', db_host,
            '-p', str(db_port),
            '-U', db_user,
            '-d', db_name,
            '-c',  # Clean (drop objects before recreating)
            '-F', 'c',  # Custom format
            backup_path
        ]

        try:
            result = subprocess.run(cmd, env=env, capture_output=True, text=True, check=True)
            print(f"✅ Yedek geri yüklendi: {backup_path}")
            return True
        except subprocess.CalledProcessError as e:
            print(f"❌ Yedek geri yükleme hatası: {e.stderr}")
            raise

    def export_all_tables_to_excel(self, db: Session) -> str:
        """Tüm tabloları Excel formatında export et"""
        from app.models import (
            User, Territory, Dealer, Posm, PosmTransfer, 
            Request, Photo, Depot, AuditLog, ScheduledReport
        )
        from app.models.user import user_depots
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        excel_filename = f"backup_excel_all_tables_{timestamp}.xlsx"
        excel_path = self.backup_dir / excel_filename
        
        wb = Workbook()
        wb.remove(wb.active)  # Varsayılan sheet'i kaldır
        
        # Tablo modelleri ve isimleri
        tables_config = [
            ("Kullanıcılar", User, self._user_to_dict),
            ("Depolar", Depot, self._depot_to_dict),
            ("Bölgeler", Territory, self._territory_to_dict),
            ("Bayiler", Dealer, self._dealer_to_dict),
            ("POSM", Posm, self._posm_to_dict),
            ("POSM Transferler", PosmTransfer, self._posm_transfer_to_dict),
            ("Talepler", Request, self._request_to_dict),
            ("Fotoğraflar", Photo, self._photo_to_dict),
            ("Audit Loglar", AuditLog, self._audit_log_to_dict),
            ("Zamanlanmış Raporlar", ScheduledReport, self._scheduled_report_to_dict),
        ]
        
        for sheet_name, model, converter_func in tables_config:
            try:
                ws = wb.create_sheet(title=sheet_name)
                
                # Tablo var mı kontrol et
                table_name = model.__tablename__
                from sqlalchemy import inspect as sql_inspect
                inspector = sql_inspect(db.bind)
                if table_name not in inspector.get_table_names():
                    print(f"⚠️ Tablo '{table_name}' bulunamadı, atlanıyor...")
                    wb.remove(ws)  # Oluşturulan boş sheet'i kaldır
                    continue
                
                records = db.query(model).all()
                
                if not records:
                    # Boş tablo için başlık satırı ekle
                    try:
                        headers = self._get_model_headers(model)
                        ws.append(headers)
                        # Başlık stilini uygula
                        self._style_header_row(ws, 1)
                    except Exception as e:
                        print(f"⚠️ '{sheet_name}' için başlık oluşturulamadı: {e}")
                        # İlk kayıttan başlık almayı dene
                        try:
                            # Model'den direkt sütun isimlerini al
                            mapper = sql_inspect(model)
                            headers = [column.key for column in mapper.columns]
                            ws.append(headers)
                            self._style_header_row(ws, 1)
                        except:
                            ws.append(["ID", "Veri Yok"])
                            self._style_header_row(ws, 1)
                else:
                    # İlk kayıttan başlıkları al
                    try:
                        first_record = records[0]
                        headers = list(converter_func(first_record, db).keys())
                        ws.append(headers)
                        self._style_header_row(ws, 1)
                        
                        # Verileri ekle
                        for record in records:
                            try:
                                row_data = converter_func(record, db)
                                ws.append([row_data.get(h, '') for h in headers])
                            except Exception as e:
                                print(f"⚠️ '{sheet_name}' için kayıt işlenirken hata: {e}")
                                continue
                        
                        # Sütun genişliklerini ayarla
                        for col in range(1, len(headers) + 1):
                            if col <= 26:
                                col_letter = chr(64 + col)
                            else:
                                col_letter = chr(64 + (col - 1) // 26) + chr(64 + ((col - 1) % 26) + 1)
                            ws.column_dimensions[col_letter].width = 20
                    except Exception as e:
                        print(f"⚠️ '{sheet_name}' işlenirken hata: {e}")
                        ws.append(["Hata", str(e)])
                        self._style_header_row(ws, 1)
                        
            except Exception as e:
                print(f"❌ '{sheet_name}' sheet'i oluşturulamadı: {e}")
                # Sheet oluşturulduysa kaldır
                try:
                    if 'ws' in locals() and ws in wb.worksheets:
                        wb.remove(ws)
                except:
                    pass
                continue
        
        wb.save(excel_path)
        print(f"✅ Excel export oluşturuldu: {excel_path}")
        return str(excel_path)
    
    def _get_model_headers(self, model):
        """Model sütunlarını al"""
        mapper = inspect(model)
        return [column.key for column in mapper.columns]
    
    def _style_header_row(self, ws, row_num):
        """Başlık satırını stilize et"""
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        for cell in ws[row_num]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
    
    def _user_to_dict(self, user: Any, db: Session) -> Dict:
        """User modelini dict'e çevir"""
        depot_names = [d.name for d in user.depots] if user.depots else []
        return {
            "ID": user.id,
            "Ad": user.name,
            "E-posta": user.email,
            "Rol": user.role,
            "Depo (Eski)": user.depot.name if user.depot else "",
            "Depolar": ", ".join(depot_names),
            "Oluşturulma": user.created_at.strftime("%Y-%m-%d %H:%M:%S") if user.created_at else "",
            "Güncellenme": user.updated_at.strftime("%Y-%m-%d %H:%M:%S") if user.updated_at else "",
        }
    
    def _depot_to_dict(self, depot: Any, db: Session) -> Dict:
        """Depot modelini dict'e çevir"""
        return {
            "ID": depot.id,
            "Ad": depot.name,
            "Kod": depot.code,
        }
    
    def _territory_to_dict(self, territory: Any, db: Session) -> Dict:
        """Territory modelini dict'e çevir"""
        return {
            "ID": territory.id,
            "Ad": territory.name,
        }
    
    def _dealer_to_dict(self, dealer: Any, db: Session) -> Dict:
        """Dealer modelini dict'e çevir"""
        return {
            "ID": dealer.id,
            "Kod": dealer.code,
            "Ad": dealer.name,
            "Bölge": dealer.territory.name if dealer.territory else "",
            "Depo": dealer.depot.name if dealer.depot else "",
            "Enlem": float(dealer.latitude) if dealer.latitude else "",
            "Boylam": float(dealer.longitude) if dealer.longitude else "",
        }
    
    def _posm_to_dict(self, posm: Any, db: Session) -> Dict:
        """Posm modelini dict'e çevir"""
        return {
            "ID": posm.id,
            "Ad": posm.name,
            "Depo": posm.depot.name if posm.depot else "",
            "Hazır Miktar": posm.ready_count,
            "Onarım Bekleyen": posm.repair_pending_count,
            "Oluşturulma": posm.created_at.strftime("%Y-%m-%d %H:%M:%S") if posm.created_at else "",
            "Güncellenme": posm.updated_at.strftime("%Y-%m-%d %H:%M:%S") if posm.updated_at else "",
        }
    
    def _posm_transfer_to_dict(self, transfer: Any, db: Session) -> Dict:
        """PosmTransfer modelini dict'e çevir"""
        return {
            "ID": transfer.id,
            "POSM": transfer.posm.name if transfer.posm else "",
            "Kaynak Depo": transfer.from_depot.name if transfer.from_depot else "",
            "Hedef Depo": transfer.to_depot.name if transfer.to_depot else "",
            "Miktar": transfer.quantity,
            "Tip": transfer.transfer_type,
            "Notlar": transfer.notes or "",
            "Transfer Eden": transfer.transferred_by_user.name if transfer.transferred_by_user else "",
            "Tarih": transfer.created_at.strftime("%Y-%m-%d %H:%M:%S") if transfer.created_at else "",
        }
    
    def _request_to_dict(self, request: Any, db: Session) -> Dict:
        """Request modelini dict'e çevir"""
        return {
            "ID": request.id,
            "Kullanıcı": request.user.name if request.user else "",
            "Bayi": request.dealer.name if request.dealer else "",
            "Bölge": request.territory.name if request.territory else "",
            "Depo": request.depot.name if request.depot else "",
            "Mevcut POSM": request.current_posm or "",
            "İş Tipi": request.job_type,
            "İş Detayı": request.job_detail or "",
            "Talep Tarihi": request.request_date.strftime("%Y-%m-%d %H:%M:%S") if request.request_date else "",
            "İstenen Tarih": request.requested_date.strftime("%Y-%m-%d") if request.requested_date else "",
            "Planlanan Tarih": request.planned_date.strftime("%Y-%m-%d") if request.planned_date else "",
            "POSM": request.posm.name if request.posm else "",
            "Durum": request.status,
            "Öncelik": request.priority,
            "Tamamlanma Açıklaması": request.job_done_desc or "",
            "Tamamlanma Tarihi": request.completed_date.strftime("%Y-%m-%d") if request.completed_date else "",
            "Tamamlayan": request.completed_by_user.name if request.completed_by_user else "",
            "Enlem": float(request.latitude) if request.latitude else "",
            "Boylam": float(request.longitude) if request.longitude else "",
            "Güncellenme": request.updated_at.strftime("%Y-%m-%d %H:%M:%S") if request.updated_at else "",
        }
    
    def _photo_to_dict(self, photo: Any, db: Session) -> Dict:
        """Photo modelini dict'e çevir"""
        return {
            "ID": photo.id,
            "Talep ID": photo.request_id,
            "Dosya Adı": photo.file_name,
            "Yol/URL": photo.path_or_url,
            "MIME Tipi": photo.mime_type or "",
            "Oluşturulma": photo.created_at.strftime("%Y-%m-%d %H:%M:%S") if photo.created_at else "",
        }
    
    def _audit_log_to_dict(self, log: Any, db: Session) -> Dict:
        """AuditLog modelini dict'e çevir"""
        return {
            "ID": log.id,
            "Kullanıcı": log.user.name if log.user else "Sistem",
            "Kullanıcı E-posta": log.user.email if log.user else "",
            "Eylem": log.action,
            "Varlık Tipi": log.entity_type,
            "Varlık ID": log.entity_id or "",
            "Açıklama": log.description or "",
            "Eski Değerler": str(log.old_values) if log.old_values else "",
            "Yeni Değerler": str(log.new_values) if log.new_values else "",
            "IP Adresi": log.ip_address or "",
            "User Agent": log.user_agent or "",
            "Tarih": log.created_at.strftime("%Y-%m-%d %H:%M:%S") if log.created_at else "",
        }
    
    def _scheduled_report_to_dict(self, report: Any, db: Session) -> Dict:
        """ScheduledReport modelini dict'e çevir"""
        depot_names = []
        if report.depot_ids:
            from app.models.depot import Depot
            depots = db.query(Depot).filter(Depot.id.in_(report.depot_ids)).all()
            depot_names = [d.name for d in depots]
        
        user_names = []
        if report.recipient_user_ids:
            from app.models.user import User
            users = db.query(User).filter(User.id.in_(report.recipient_user_ids)).all()
            user_names = [u.name for u in users]
        
        return {
            "ID": report.id,
            "Ad": report.name,
            "Rapor Tipi": report.report_type,
            "Cron İfadesi": report.cron_expression,
            "Aktif": "Evet" if report.is_active else "Hayır",
            "Depolar": ", ".join(depot_names) if depot_names else "Tümü",
            "Alıcılar": ", ".join(user_names),
            "Durum Filtresi": ", ".join(report.status_filter) if report.status_filter else "",
            "İş Tipi Filtresi": ", ".join(report.job_type_filter) if report.job_type_filter else "",
            "Son Gönderim": report.last_sent_at.strftime("%Y-%m-%d %H:%M:%S") if report.last_sent_at else "",
            "Sonraki Çalışma": report.next_run_at.strftime("%Y-%m-%d %H:%M:%S") if report.next_run_at else "",
            "Oluşturulma": report.created_at.strftime("%Y-%m-%d %H:%M:%S") if report.created_at else "",
        }

    def create_full_system_backup(self, db: Session, db_host: str, db_port: int, db_name: str, db_user: str, db_password: str) -> str:
        """Sistemin komple yedeğini oluştur (DB + Excel + Uploads + Config)"""
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        backup_name = f"full_system_backup_{timestamp}"
        backup_dir = self.backup_dir / backup_name
        backup_dir.mkdir(parents=True, exist_ok=True)
        
        try:
            # 1. PostgreSQL yedeği
            print("📦 PostgreSQL yedeği oluşturuluyor...")
            db_backup_path = self.create_database_backup(db_host, db_port, db_name, db_user, db_password)
            db_backup_filename = os.path.basename(db_backup_path)
            shutil.copy2(db_backup_path, backup_dir / db_backup_filename)
            
            # 2. Excel export
            print("📊 Excel export oluşturuluyor...")
            excel_path = self.export_all_tables_to_excel(db)
            excel_filename = os.path.basename(excel_path)
            shutil.copy2(excel_path, backup_dir / excel_filename)
            
            # 3. Uploads klasörünü kopyala
            print("📁 Uploads klasörü kopyalanıyor...")
            uploads_dir = Path(settings.UPLOAD_DIR if hasattr(settings, 'UPLOAD_DIR') else './uploads')
            if uploads_dir.exists():
                shutil.copytree(uploads_dir, backup_dir / 'uploads', dirs_exist_ok=True)
            
            # 4. Config dosyalarını kopyala (varsa)
            print("⚙️ Config dosyaları kopyalanıyor...")
            config_files = ['.env', 'docker-compose.yml']
            for config_file in config_files:
                config_path = Path(config_file)
                if config_path.exists():
                    shutil.copy2(config_path, backup_dir / config_path.name)
            
            # 5. Backup bilgileri dosyası oluştur
            info_file = backup_dir / 'backup_info.txt'
            with open(info_file, 'w', encoding='utf-8') as f:
                f.write(f"Sistem Yedeği Bilgileri\n")
                f.write(f"{'='*50}\n\n")
                f.write(f"Oluşturulma Tarihi: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n")
                f.write(f"Veritabanı: {db_name}\n")
                f.write(f"PostgreSQL Yedeği: {db_backup_filename}\n")
                f.write(f"Excel Export: {excel_filename}\n")
                f.write(f"Uploads Klasörü: {'Mevcut' if uploads_dir.exists() else 'Yok'}\n")
                f.write(f"\nİçerik:\n")
                f.write(f"- PostgreSQL veritabanı yedeği (.sql)\n")
                f.write(f"- Tüm tablolar Excel formatında (.xlsx)\n")
                f.write(f"- Uploads klasörü (fotoğraflar ve diğer dosyalar)\n")
                f.write(f"- Config dosyaları (.env, docker-compose.yml)\n")
            
            # 6. Tüm dosyaları ZIP olarak arşivle
            print("🗜️ ZIP arşivi oluşturuluyor...")
            zip_filename = f"{backup_name}.zip"
            zip_path = self.backup_dir / zip_filename
            
            with zipfile.ZipFile(zip_path, 'w', zipfile.ZIP_DEFLATED) as zipf:
                for root, dirs, files in os.walk(backup_dir):
                    for file in files:
                        file_path = Path(root) / file
                        arcname = file_path.relative_to(backup_dir)
                        zipf.write(file_path, arcname)
            
            # 7. Geçici klasörü sil
            shutil.rmtree(backup_dir)
            
            print(f"✅ Sistem yedeği oluşturuldu: {zip_path}")
            return str(zip_path)
            
        except Exception as e:
            # Hata durumunda geçici klasörü temizle
            if backup_dir.exists():
                shutil.rmtree(backup_dir, ignore_errors=True)
            print(f"❌ Sistem yedeği oluşturma hatası: {e}")
            raise
